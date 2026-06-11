#!/usr/bin/env python3
"""Generate the Box Office Terminals workbook (structure + formulas).

Builds an .xlsx with the sheets, structured tables, dropdowns and formula
columns that replace the reporting-dashboard box-office feature. Design choices
agreed for the Excel version:

  * ONE terminal per hire row (a multi-terminal loan = multiple rows). This
    keeps the double-booking check, utilisation and derived status clean.
  * NO Gantt/Schedule sheet — dropped as the weakest fit for Excel.
  * Single-owner editing — no concurrency handling needed.

Most "derived" values (terminal status, available-from, utilisation) are plain
worksheet formulas defined here. The behavioural rules that a formula can't
express — rejecting a double-booking, completing the prior hire on a
continuation, the trial/payment guard — live in VBA (box_office_excel/vba/*.bas)
and are pasted in once (see SETUP.md).

Run (hand to operator — writes a local file only, no regulated data):
    python3 box_office_excel/build_workbook.py
    # -> box_office_excel/BoxOfficeTerminals.xlsx

British spellings throughout (per project convention).
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Written as .xlsx (no macros yet). openpyxl can't author VBA, and saving a
# macro-declared .xlsm with no vbaProject.bin makes Excel (esp. on Mac) reject
# the file as corrupt. The macros are added in Excel per SETUP.md, and Excel
# writes a valid .xlsm when you then Save As macro-enabled.
OUT = "box_office_excel/BoxOfficeTerminals.xlsx"
DATA_ROWS = 500

# ── domain constants ────────────────────────────────────────────────────────
# Cells store the HUMAN-READABLE label (no machine values / underscores).
# The VBA + migration work in these same labels.
HIRE_STATUSES = [
    "Draft", "Pending payment", "Confirmed", "Shipped",
    "In use", "Returned", "Completed", "Cancelled",
]
ACTIVE_STATUSES = ["Confirmed", "Shipped", "In use"]      # committed / out
DELIVERY_METHODS = ["Ship", "Drop off", "Collect"]
TERMINAL_MODELS = ["BBPOS WisePOS E"]

# DB machine-value -> display label, for the migration.
STATUS_FROM_DB = {
    "draft": "Draft", "pending_payment": "Pending payment", "confirmed": "Confirmed",
    "shipped": "Shipped", "in_use": "In use", "returned": "Returned",
    "completed": "Completed", "cancelled": "Cancelled",
}
DELIVERY_FROM_DB = {"ship": "Ship", "drop_off": "Drop off", "collect": "Collect"}

# Columns carry a friendly HEADER (what users see, and what VBA/migration
# reference by name), a HIDDEN flag (internal IDs the macros need but users
# shouldn't see), and a KIND that drives formatting/validation:
#   id|text|status|tick|date|terminal|money|int|delivery
# One terminal per hire row.
HIRE_COLUMNS = [
    ("Ref",                  True,  "id"),       # was HireId
    ("Account ref",          True,  "id"),       # was AccountId
    ("Account",              False, "text"),
    ("Contact name",         False, "text"),
    ("Contact email",        False, "text"),
    ("Contact phone",        False, "text"),
    ("Status",               False, "status"),
    ("Trial",                False, "tick"),
    ("Hire from",            False, "date"),
    ("Hire to",              False, "date"),
    ("Terminal",             False, "terminal"),
    ("Cradles",              False, "int"),
    ("Send out by",          False, "delivery"),
    ("Return by",            False, "delivery"),
    ("Shipping address",     False, "text"),
    ("Paid",                 False, "tick"),
    ("Amount due",           False, "money"),
    ("Payment ref",          False, "text"),
    ("Continues from (ref)", True,  "id"),       # was ContinuesFromHireId
    ("Box Office Web",       False, "tick"),
    ("Terminals linked",     False, "tick"),
    ("Notes",                False, "text"),
    ("Changed by",           False, "text"),
    ("Changed at",           False, "text"),
]

TERMINAL_COLUMNS = [
    ("Terminal",        False, "text"),    # the terminal ID (human-named e.g. "01")
    ("Model",           False, "model"),
    ("Status",          False, "formula"),
    ("Available from",  False, "formula"),
    ("Future hires",    False, "formula"),
    ("Utilisation",     False, "formula"),
    ("Notes",           False, "text"),
]

# Friendly-header → internal key, so the migration + VBA stay readable.
HIRE_HEADERS = [c[0] for c in HIRE_COLUMNS]
TERMINAL_HEADERS = [c[0] for c in TERMINAL_COLUMNS]

HEADER_FILL = PatternFill("solid", fgColor="2F5233")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def make_table(ws, name, ncols, nrows):
    last_col = get_column_letter(ncols)
    tbl = Table(displayName=name, ref=f"A1:{last_col}{nrows + 1}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)


def add_list_validation(ws, col_letter, first_row, last_row, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.error = "Pick a value from the list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


# Tick columns hold TRUE/FALSE. Native Excel checkboxes (Insert -> Checkbox,
# added in Microsoft 365) sit on these cells and toggle the value; until they're
# inserted the cells just show TRUE/FALSE. openpyxl can't author the checkbox
# control, so inserting it is a one-time manual step (see SETUP.md).


def build():
    wb = Workbook()

    # ── Lists (dropdown sources; hidden) ────────────────────────────────────
    lists = wb.active
    lists.title = "Lists"
    cols = [("A", "Statuses", HIRE_STATUSES),
            ("B", "DeliveryMethods", DELIVERY_METHODS),
            ("C", "Models", TERMINAL_MODELS),
            ("D", "Bool", ["TRUE", "FALSE"])]
    for letter, head, vals in cols:
        lists[f"{letter}1"] = head
        lists[f"{letter}1"].font = Font(bold=True)
        for i, v in enumerate(vals, start=2):
            lists[f"{letter}{i}"] = v
    lists.sheet_state = "hidden"

    last = DATA_ROWS + 1  # last data row

    # ── Terminals first (so the Terminal-id named range exists for the dropdown)
    terms = wb.create_sheet("Terminals")
    for i, (hdr, _, _) in enumerate(TERMINAL_COLUMNS, start=1):
        terms.cell(row=1, column=i, value=hdr)
    style_header(terms, len(TERMINAL_COLUMNS))
    make_table(terms, "TerminalsTable", len(TERMINAL_COLUMNS), DATA_ROWS)
    tc = {hdr: get_column_letter(i + 1) for i, (hdr, _, _) in enumerate(TERMINAL_COLUMNS)}

    add_list_validation(terms, tc["Model"], 2, last,
                        f"=Lists!$C$2:$C${len(TERMINAL_MODELS) + 1}")

    # Named range over the Terminal-id column (data validation accepts a name,
    # but NOT a structured table reference — that was what corrupted the file).
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName(
        "TerminalIds", attr_text=f"Terminals!${tc['Terminal']}$2:${tc['Terminal']}${last}"))

    # Status / Available from / Future hires / Utilisation are filled by the
    # RefreshTerminals macro (array formulas were fragile on Mac Excel). Just set
    # display formats here; the cells start blank.
    for r in range(2, last + 1):
        terms[f"{tc['Available from']}{r}"].number_format = "dd mmm yyyy"
        terms[f"{tc['Utilisation']}{r}"].number_format = "0%"
    terms.column_dimensions[tc["Terminal"]].width = 12
    terms.column_dimensions[tc["Notes"]].width = 28
    terms.column_dimensions[tc["Available from"]].width = 14
    terms.column_dimensions[tc["Utilisation"]].width = 12

    # ── Hires (source of truth) ─────────────────────────────────────────────
    hires = wb.create_sheet("Hires")
    for i, (hdr, _, _) in enumerate(HIRE_COLUMNS, start=1):
        hires.cell(row=1, column=i, value=hdr)
    style_header(hires, len(HIRE_COLUMNS))
    make_table(hires, "HiresTable", len(HIRE_COLUMNS), DATA_ROWS)
    hc = {hdr: get_column_letter(i + 1) for i, (hdr, _, _) in enumerate(HIRE_COLUMNS)}

    add_list_validation(hires, hc["Status"], 2, last,
                        f"=Lists!$A$2:$A${len(HIRE_STATUSES) + 1}")
    for f in ("Send out by", "Return by"):
        add_list_validation(hires, hc[f], 2, last,
                            f"=Lists!$B$2:$B${len(DELIVERY_METHODS) + 1}")
    for f in ("Trial", "Paid", "Box Office Web", "Terminals linked"):
        tick_dv(hires, hc[f], 2, last)
    # Terminal dropdown via the named range (Excel-safe).
    add_list_validation(hires, hc["Terminal"], 2, last, "=TerminalIds")

    for f in ("Hire from", "Hire to"):
        for r in range(2, last + 1):
            hires[f"{hc[f]}{r}"].number_format = "dd mmm yyyy"
    for r in range(2, last + 1):
        hires[f"{hc['Amount due']}{r}"].number_format = "£#,##0.00"

    for hdr, w in {"Account": 24, "Contact name": 18, "Contact email": 24,
                   "Shipping address": 28, "Notes": 28, "Changed at": 18,
                   "Trial": 6, "Paid": 6, "Cradles": 8}.items():
        hires.column_dimensions[hc[hdr]].width = w

    # Hide the internal-id columns on both sheets.
    for hdr, hidden, _ in HIRE_COLUMNS:
        if hidden:
            hires.column_dimensions[hc[hdr]].hidden = True

    # ── ReadMe ──────────────────────────────────────────────────────────────
    info = wb.create_sheet("ReadMe", 0)
    msg = [
        "Box Office Terminals — Excel workbook",
        "",
        "Replaces the reporting-dashboard box-office feature.",
        "  • Hires sheet = source of truth (ONE terminal per row).",
        "  • Terminals sheet = your kit; Status / Utilisation / Available are formulas.",
        "",
        "Tick columns (Trial, Paid, …) are checkboxes — tick = yes. (Insert the",
        "checkbox controls once per SETUP.md; they store TRUE/FALSE.)",
        "Hidden reference columns (Ref / Account ref) are used by the macros — leave them.",
        "",
        "Business rules that formulas can't enforce live in VBA macros — paste them",
        "in once (see SETUP.md). They are:",
        "  • Double-booking guard — blocks a terminal on two overlapping active hires.",
        "  • Continuation — completes the prior hire, carries contact/address/dates.",
        "  • Trial bypasses the payment requirement.",
        "Until the macros are added this is a structured spreadsheet without the guards.",
        "",
        "Single-owner editing only — do not have two people editing at once.",
    ]
    for i, line in enumerate(msg, start=1):
        info.cell(row=i, column=1, value=line)
    info.column_dimensions["A"].width = 84
    info["A1"].font = Font(bold=True, size=14)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print("Next: paste the .bas modules (box_office_excel/vba/) per SETUP.md,")
    print("then Save As from Excel as .xlsm to keep the macros.")


def tick_dv(ws, col_letter, first_row, last_row):
    """Seed a checkbox column to FALSE so the native checkbox renders unchecked.
    (No data validation — the inserted checkbox control owns the cell.)"""
    for r in range(first_row, last_row + 1):
        ws[f"{col_letter}{r}"] = False
        ws[f"{col_letter}{r}"].alignment = Alignment(horizontal="center")




if __name__ == "__main__":
    build()
