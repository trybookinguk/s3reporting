#!/usr/bin/env python3
"""Migrate existing box-office data from box_office.db into the Excel workbook.

Reads the live SQLite store (box_office_hires / box_office_inventory /
box_office_hire_items / box_office_settings) and writes the rows into
BoxOfficeTerminals.xlsx — into the Hires and Terminals structured tables.

Model change: the app allows multiple terminals per hire (a join table); the
Excel workbook is ONE terminal per row. A hire holding N terminals therefore
becomes N rows here (sharing the same HireId base, suffixed -1, -2, … so they
stay distinct). A hire with no terminals becomes a single row with Terminal blank.

DATA-PROTECTION NOTE: this reads regulated box-office data, so per project
convention it is run by the human operator ON THE PI, not by the assistant.

Run ON THE PI (where box_office.db lives), after build_workbook.py has produced
the .xlsm:
    cd /root/s3reporting
    python3 box_office_excel/migrate_from_db.py \
        --db /root/s3reporting/.cache/prepared/box_office.db \
        --xlsm box_office_excel/BoxOfficeTerminals.xlsx

Idempotent-ish: it rewrites the data rows each run (clears prior data first).
"""

import argparse
import sqlite3
from openpyxl import load_workbook

# DB machine-value -> display label (must match build_workbook.py).
STATUS_FROM_DB = {
    "draft": "Draft", "pending_payment": "Pending payment", "confirmed": "Confirmed",
    "shipped": "Shipped", "in_use": "In use", "returned": "Returned",
    "completed": "Completed", "cancelled": "Cancelled",
}
DELIVERY_FROM_DB = {"ship": "Ship", "drop_off": "Drop off", "collect": "Collect"}

# Friendly headers — must match build_workbook.py exactly (display names).
HIRE_HEADERS = [
    "Ref", "Account ref", "Account", "Contact name", "Contact email",
    "Contact phone", "Status", "Trial", "Hire from", "Hire to",
    "Terminal", "Cradles", "Send out by", "Return by", "Shipping address",
    "Paid", "Amount due", "Payment ref", "Continues from (ref)",
    "Box Office Web", "Terminals linked", "Notes", "Changed by", "Changed at",
]
TERMINAL_HEADERS = [
    "Terminal", "Model", "Status", "Available from",
    "Future hires", "Utilisation", "Notes",
]
# Terminal-sheet columns that are formulas (don't overwrite on migrate).
TERMINAL_FORMULA_HEADERS = {"Status", "Available from", "Future hires", "Utilisation"}


def iso_date(v):
    """SQLite stores dates as ISO strings; return YYYY-MM-DD or None."""
    if not v:
        return None
    return str(v)[:10]


def fmt_address(json_str):
    """Flatten the shipping_address JSON blob to a single line."""
    import json
    try:
        a = json.loads(json_str) if json_str else {}
    except Exception:
        return json_str or ""
    parts = [a.get("line1"), a.get("line2"), a.get("city"), a.get("postcode"), a.get("country")]
    return ", ".join(p for p in parts if p)


def load_db(db_path):
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    hires = [dict(r) for r in cur.execute("SELECT * FROM box_office_hires")]
    inventory = [dict(r) for r in cur.execute(
        "SELECT * FROM box_office_inventory WHERE type='terminal'")]
    items = [dict(r) for r in cur.execute("SELECT * FROM box_office_hire_items")]
    settings = {r["key"]: r["value"] for r in cur.execute("SELECT * FROM box_office_settings")}
    con.close()

    # hire_id -> [terminal item_ids]
    by_hire = {}
    for it in items:
        by_hire.setdefault(it["hire_id"], []).append(it["item_id"])
    return hires, inventory, by_hire, settings


def hire_to_rows(h, terminal_ids):
    """Expand one DB hire into one workbook row per terminal (≥1 rows)."""
    base = {
        "Ref": h["id"],
        "Account ref": h["account_id"],
        "Account": h["account_name"],
        "Contact name": h["contact_name"],
        "Contact email": h["contact_email"],
        "Contact phone": h.get("contact_phone") or "",
        "Status": STATUS_FROM_DB.get(h["status"], h["status"]),
        "Trial": bool(h.get("is_trial")),
        "Hire from": iso_date(h["hire_from"]),
        "Hire to": iso_date(h["hire_to"]),
        "Terminal": "",
        "Cradles": h.get("cradle_count") or 0,
        "Send out by": DELIVERY_FROM_DB.get(h["outbound_method"], h["outbound_method"]),
        "Return by": DELIVERY_FROM_DB.get(h["return_method"], h["return_method"]),
        "Shipping address": fmt_address(h.get("shipping_address")),
        "Paid": bool(h.get("payment_received")),
        "Amount due": (h["amount_due_pence"] / 100) if h.get("amount_due_pence") is not None else None,
        "Payment ref": h.get("payment_reference") or "",
        "Continues from (ref)": h.get("continues_from_hire_id") or "",
        "Box Office Web": bool(h.get("box_office_web_enabled")),
        "Terminals linked": bool(h.get("terminals_linked_to_account")),
        "Notes": h.get("notes") or "",
        "Changed by": h.get("changed_by") or "",
        "Changed at": iso_date(h.get("changed_at")),
    }
    if not terminal_ids:
        return [base]
    rows = []
    multi = len(terminal_ids) > 1
    for i, tid in enumerate(sorted(terminal_ids), start=1):
        row = dict(base)
        row["Terminal"] = tid
        # Keep Refs unique across split rows; continuation links still resolve to
        # the FIRST row's ref (which equals the original DB id).
        if multi and i > 1:
            row["Ref"] = f"{h['id']}-{i}"
        rows.append(row)
    return rows


def clear_table_body(ws, ncols, keep_formula_cols=None):
    """Blank out data rows (row 2..) for the given columns, leaving formulas."""
    keep = keep_formula_cols or set()
    for r in range(2, ws.max_row + 1):
        for c in range(1, ncols + 1):
            if c in keep:
                continue
            ws.cell(row=r, column=c).value = None


def write_rows(ws, columns, dict_rows):
    # Booleans are written as real TRUE/FALSE so the native checkboxes bind to them.
    idx = {name: i + 1 for i, name in enumerate(columns)}
    for ri, d in enumerate(dict_rows, start=2):
        for name, col in idx.items():
            ws.cell(row=ri, column=col, value=d.get(name))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--xlsm", required=True)
    args = ap.parse_args()

    hires, inventory, by_hire, settings = load_db(args.db)

    # Build hire rows (expanded one-terminal-per-row).
    hire_rows = []
    for h in sorted(hires, key=lambda x: (x["hire_from"] or "", x["id"])):
        hire_rows.extend(hire_to_rows(h, by_hire.get(h["id"], [])))

    # Terminal rows (derived columns are formulas — only fill the inputs).
    term_rows = []
    for inv in sorted(inventory, key=lambda x: x["id"]):
        term_rows.append({
            "Terminal": inv["id"],
            "Model": inv.get("model") or "",
            "Notes": inv.get("notes") or "",
        })

    wb = load_workbook(args.xlsm)
    hires_ws = wb["Hires"]
    terms_ws = wb["Terminals"]

    # Clear existing data bodies (Hires has no formulas; Terminals keeps its
    # formula columns).
    clear_table_body(hires_ws, len(HIRE_HEADERS))
    term_formula_cols = {TERMINAL_HEADERS.index(c) + 1 for c in TERMINAL_FORMULA_HEADERS}
    clear_table_body(terms_ws, len(TERMINAL_HEADERS), keep_formula_cols=term_formula_cols)

    write_rows(hires_ws, HIRE_HEADERS, hire_rows)
    # Terminals: write only the input columns by header (skip formula columns).
    t_idx = {hdr: i + 1 for i, hdr in enumerate(TERMINAL_HEADERS)}
    for ri, d in enumerate(term_rows, start=2):
        for hdr, ci in t_idx.items():
            if hdr in d:
                terms_ws.cell(row=ri, column=ci, value=d[hdr])

    # Seed cradles-owned into the ReadMe (a known cell) for reference.
    if "cradles_owned" in settings:
        wb["ReadMe"].cell(row=20, column=1,
                          value=f"Cradles owned (migrated): {settings['cradles_owned']}")

    wb.save(args.xlsm)
    print(f"Migrated {len(hires)} hires -> {len(hire_rows)} rows, "
          f"{len(inventory)} terminals into {args.xlsm}")
    if settings.get("cradles_owned"):
        print(f"Cradles owned: {settings['cradles_owned']}")


if __name__ == "__main__":
    main()
